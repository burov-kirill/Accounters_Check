import openpyxl

from openpyxl.chart.axis import ChartLines
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.marker import DataPoint
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.formatting import Rule
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment, DEFAULT_FONT
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import ColorScale, FormatObject
from openpyxl.styles import Color
from openpyxl.chart import BarChart, Reference, BarChart3D, DoughnutChart
from openpyxl.chart.layout import Layout, ManualLayout
import datetime

PLOT_DICT = {'main_table': '',
             'summary_percent': 'Общее выполнение плана по срокам закрытия, %',
             'superior_percent': 'Доля компаний, нарушающих регламентированные',
             'date_table': 'Распределение дат итогового закрытия отчётного месяца',
             'term_data': 'Структурный анализ итогового закрытия отчётного месяца',
             'plan_table': 'Структурный анализ соблюдения Регламентируемых сроков',
             'best_company': 'ТОП-10 «Лидеры закрытия»',
             'worst_company': 'ТОП-10 «Зона роста»',
             'average_table': 'Рейтинг закрытия участков РСБУ на основе балльной системы'}
def export_to_excel(file, month, save_path):
    wb = openpyxl.Workbook()
    result_sheet = wb.create_sheet(month)
    del wb[wb.get_sheet_names()[0]]
    decoration_table(result_sheet, file.report, 2, file.tables['main_table'])
    name = 'Отчет по закрытию.xlsx'
    save_path = fr'{save_path}\{name}'
    _font = Font(name="Arial Narrow", sz=9, b=False)
    {k: setattr(DEFAULT_FONT, k, v) for k, v in _font.__dict__.items()}
    tables_sheet = wb.create_sheet('Таблицы')
    init_col = 2
    row = 2
    max_row = 0
    for number, (key, value) in enumerate(file.tables.items(), 3):
        if key != 'main_table':
            if key in ('summary_percent', 'superior_percent','term_data',
                                       'plan_table', 'best_company', 'worst_company'):
                paste_table(tables_sheet, value, init_col, number, row, True, True)
            else:
                paste_table(tables_sheet, value, init_col, number, row, True)
            if key in ('summary_percent', 'superior_percent'):
                chart1 = BarChart()
                chart1.type = "col"
                chart1.style = 10
                chart1.title = PLOT_DICT[key]
                chart1.title.buClrTx = '007bfb'
                data = Reference(tables_sheet, min_col=3, min_row=row, max_row=row+len(value), max_col=3)
                cats = Reference(tables_sheet, min_col=2, min_row=row+1, max_row=row+len(value))
                chart1.add_data(data, titles_from_data=True)
                chart1.set_categories(cats)
                # chart1.shape = 4
                chart1.y_axis.number_format = '0%'
                chart1.y_axis.delete = True
                chart1.x_axis.majorGridlines = None
                chart1.y_axis.majorGridlines = None
                chart1.legend = None
                slices = [DataPoint(idx=i) for i in range(len(value))]
                color_list = ["CFCFCF", "007bfb", "ffffff"]
                for idx, point in enumerate(slices):
                    col_idx = idx % len(color_list)
                    point.graphicalProperties.solidFill = color_list[col_idx]
                    point.graphicalProperties.ln.solidFill = color_list[col_idx]
                    if col_idx == 2:
                        point.graphicalProperties.ln.solidFill = color_list[1]
                        point.graphicalProperties.ln.w = 2
                chart1.series[0].data_points = slices
                # chart1.y_axis.majorGridlines.spPr = GraphicalProperties(noFill='True')
                # chart1.y_axis.majorGridlines.spPr.ln = LineProperties(solidFill='000000')

                # chart1.x_axis.majorGridlines = ChartLines()
                # chart1.x_axis.majorGridlines.spPr = GraphicalProperties(noFill='True')
                # chart1.x_axis.majorGridlines.spPr.ln = LineProperties(solidFill='000000')
                chart1.dLbls = DataLabelList()
                chart1.dLbls.showVal = 1
                tables_sheet.add_chart(chart1, f"{get_column_letter(init_col + 4)}{row}")
            elif key in ('date_table','best_company', 'worst_company'):
                data = Reference(tables_sheet, min_col=3, min_row=row, max_col=3, max_row=row + len(value))
                if key == 'date_table':
                    titles = Reference(tables_sheet, min_col=4, min_row=row+1, max_row=row + len(value))
                else:
                    titles = Reference(tables_sheet, min_col=2, min_row=row + 1, max_row=row + len(value))
                chart = BarChart3D()
                chart.type = "bar"
                chart.title = PLOT_DICT[key]
                chart.add_data(data=data, titles_from_data=True)
                chart.set_categories(titles)
                chart.x_axis.majorGridlines = None
                chart.y_axis.majorGridlines = None
                chart.legend = None
                chart.dLbls = DataLabelList()
                chart.dLbls.showVal = 1
                slices = [DataPoint(idx=i) for i in range(len(value))]
                if key == 'date_table':
                    chart.y_axis.title = 'Количество ЮЛ'
                    chart.x_axis.title = 'Дата итогового закрытия'
                    color_list = ["00FF00", "ffff00", "FF0000"]
                    for idx, (point, date) in enumerate(zip(slices, value['Дата итогового закрытия'])):
                        if type(date) == datetime.date:
                            diff = file.DATE_DICT[file.LAST_COLUMN] - date
                            if diff.days == 0:
                                col_idx = 0
                            elif diff.days > 0:
                                col_idx = 1
                            else:
                                col_idx = 2
                        else:
                            col_idx = 2
                        point.graphicalProperties.solidFill = color_list[col_idx]
                        point.graphicalProperties.ln.solidFill = color_list[col_idx]
                else:
                    chart.height = 10  # default is 7.5
                    chart.width = 20
                    if key == 'best_company':
                        color_list = ["FF0000"]
                    else:
                        color_list = ['00FF00']
                    for idx, point in enumerate(slices):
                        col_idx = idx % len(color_list)
                        point.graphicalProperties.solidFill = color_list[col_idx]
                        point.graphicalProperties.ln.solidFill = color_list[col_idx]

                chart.series[0].data_points = slices
                tables_sheet.add_chart(chart, f"{get_column_letter(init_col + 4)}{row}")
            elif key in ('term_data', 'plan_table'):
                chart = DoughnutChart(holeSize=50)
                labels = Reference(tables_sheet, min_col=2, min_row=row+1, max_row=row+len(value))
                data = Reference(tables_sheet, min_col=3, min_row=row, max_row=row+len(value))
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(labels)
                chart.title = PLOT_DICT[key]
                # chart.type = "filled"
                chart.style = 1
                chart.dLbls = DataLabelList()
                chart.dLbls.showVal = 1
                slices = [DataPoint(idx=i) for i in range(len(value))]
                if key == 'term_data':
                    color_list = ["00FF00", "ffff00", "FF0000"]
                else:
                    color_list = ["ffff00", "FF0000", "00FF00"]
                for idx, point in enumerate(slices):
                    col_idx = idx % len(color_list)
                    point.graphicalProperties.solidFill = color_list[col_idx]
                    point.graphicalProperties.ln.solidFill = color_list[col_idx]
                chart.series[0].data_points = slices
                chart.legend.position = 'b'
                tables_sheet.add_chart(chart, f"{get_column_letter(init_col + 4)}{row}")
            elif key == 'average_table':
                chart = BarChart()
                chart.style = 10
                chart.title = PLOT_DICT[key]
                chart.y_axis.title = 'Средний балл закрытия'
                chart.x_axis.title = 'Участок РСБУ'
                chart.style = 11
                chart.type = "bar"
                chart.height = 25  # default is 7.5
                chart.width = 21
                chart.legend = None
                data = Reference(tables_sheet, min_col=3, min_row=row, max_row=row+len(value), max_col=3)
                cats = Reference(tables_sheet, min_col=2, min_row=row+1, max_row=row+len(value))
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(cats)
                chart.dLbls = DataLabelList()
                chart.dLbls.showVal = 1
                chart.x_axis.majorGridlines = None
                chart.y_axis.majorGridlines = None
                slices = [DataPoint(idx=i) for i in range(len(value))]
                color_list = ["00FF00", "ffff00", "FF0000"]
                for point, dt in zip(slices, value['Средний бал']):
                    if dt >= 2:
                        color_idx = 0
                    elif dt < 2 and dt > 1.5:
                        color_idx = 1
                    else:
                        color_idx = 2
                    point.graphicalProperties.solidFill = color_list[color_idx]
                    point.graphicalProperties.ln.solidFill = color_list[color_idx]
                chart.series[0].data_points = slices
                tables_sheet.add_chart(chart, f"{get_column_letter(init_col + 4)}{row}")
            row +=len(value) + 10
        # if key not in ['main_table', 'summary_percent', 'superior_percent', 'average_table']:
        #     paste_table(tables_sheet, value, init_col, number,  row, True)
        #     init_col+=len(value.columns) + 2
        #     if max_row < len(value):
        #         max_row = len(value)
        # if key == 'average_table':
        #     paste_table(tables_sheet, value, 2, number,  max_row + 5, True)
        #     init_col+=len(value.columns) + 2
    wb.save(save_path)
    wb.close()



def decoration_table(ws, table, col, other_table):
    rows = dataframe_to_rows(table, index=False)
    for r_idx, row in enumerate(rows, 2):
        for c_idx, value in enumerate(row, col):
            ws.cell(row=r_idx, column=c_idx, value=value).number_format = '#,##0.00'
            if c_idx not in(2, 3):
                ws[f"{get_column_letter(c_idx)}{r_idx}"].alignment = Alignment(horizontal="center", vertical='center',wrap_text=True)
            else:
                ws.row_dimensions[r_idx].height = 52
                if c_idx == 2:
                    ws.column_dimensions[get_column_letter(c_idx)].width = 46
                else:
                    ws.column_dimensions[get_column_letter(c_idx)].width = 21
                ws[f"{get_column_letter(c_idx)}{r_idx}"].alignment = Alignment(vertical='center', wrap_text=True)


    init_col = get_column_letter(col)
    end_col = get_column_letter(len(table.columns)+col-1)
    table_length = len(table) + 2
    tab = Table(displayName=f"Table_first",
                ref=f"{init_col}2:{end_col}{table_length}")
    style = TableStyleInfo(name='TableStyleMedium2', showFirstColumn=True,
                           showLastColumn=True, showRowStripes=False, showColumnStripes=True)
    tab.tableStyleInfo = style
    ws.add_table(tab)
    ws.row_dimensions[2].height = 67.5
    first = FormatObject(type='min')
    second = FormatObject(type='num', val=2)
    last = FormatObject(type='max')
    colors = [Color('FF0000'), Color('FFFF00'), Color('00FF00')]
    cs2 = ColorScale(cfvo=[first,second, last], color=colors)
    rule = Rule(type='colorScale', colorScale=cs2)
    for column in range(4, len(table.columns) + 2):
        ws.column_dimensions[get_column_letter(column)].width = 15
        ws.conditional_formatting.add(f'{get_column_letter(column)}3:{get_column_letter(column)}{len(table)}', rule)

    init_col = col + len(table.columns) + 3
    paste_table(ws, other_table, init_col, 2)


def paste_table(ws, table, init_col, number, init_row = 2, opt = False, is_percent = False):
    rows = dataframe_to_rows(table, index=False)
    last_col = 1
    last_row = 1
    for r_idx, row in enumerate(rows, init_row):
        last_row = r_idx
        for c_idx, value in enumerate(row, init_col):
            last_col = c_idx
            if opt:
                if c_idx != 2 and is_percent:
                    ws.cell(row=r_idx, column=c_idx, value=value).number_format = '0%'

                if c_idx == init_col:
                    ws.cell(row=r_idx, column=c_idx, value=str(value))
                elif c_idx != init_col and not is_percent:
                    ws.cell(row=r_idx, column=c_idx, value=value).number_format = '#,##0.00'
            else:
                if c_idx - init_col in (3, 4):
                    ws.cell(row=r_idx, column=c_idx, value=value).number_format = '0%'
                else:
                    ws.cell(row=r_idx, column=c_idx, value=value).number_format = '#,##0.00'
            ws[f"{get_column_letter(c_idx)}{r_idx}"].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
    tab = Table(displayName=f"Table_{number}",
                ref=f"{get_column_letter(init_col)}{init_row}:{get_column_letter(last_col)}{last_row}")
    style = TableStyleInfo(name='TableStyleMedium2', showFirstColumn=True,
                           showLastColumn=True, showRowStripes=False, showColumnStripes=True)
    tab.tableStyleInfo = style
    ws.add_table(tab)